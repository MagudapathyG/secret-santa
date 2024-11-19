from openpyxl import load_workbook, Workbook
from .employee import Employee

class FileHandler:
    @staticmethod
    def load_employees(file_path):
        employees = []
        workbook = load_workbook(file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            employees.append(Employee(name=row[0], email=row[1]))
        return employees

    @staticmethod
    def load_previous_assignments(file_path):
        assignments = {}
        workbook = load_workbook(file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            assignments[row[1]] = row[3]  # Employee_EmailID -> Secret_Child_EmailID
        return assignments

    @staticmethod
    def save_assignments(assignments, output_file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Secret Santa Assignments"

        sheet.append(["Employee_Name", "Employee_EmailID", "Secret_Child_Name", "Secret_Child_EmailID"])

        for giver, recipient in assignments:
            sheet.append([giver.name, giver.email, recipient.name, recipient.email])

        workbook.save(output_file_path)
